"""
Excel export service for fund data
"""
from typing import Dict, Any, List
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.fund import Fund
from app.models.transaction import CapitalCall, Distribution, Adjustment
from app.services.metrics_calculator import MetricsCalculator


class ExcelExporter:
    """Export fund data to Excel format"""

    def __init__(self, db: Session):
        self.db = db
        self.metrics_calculator = MetricsCalculator(db)

    def export_fund_data(self, fund_id: int) -> io.BytesIO:
        """
        Export complete fund data to Excel

        Returns:
            BytesIO object containing Excel file
        """
        # Get fund details
        fund = self.db.query(Fund).filter(Fund.id == fund_id).first()
        if not fund:
            raise ValueError(f"Fund {fund_id} not found")

        # Create workbook
        wb = Workbook()

        # Add sheets
        self._add_overview_sheet(wb, fund)
        self._add_capital_calls_sheet(wb, fund_id)
        self._add_distributions_sheet(wb, fund_id)
        self._add_adjustments_sheet(wb, fund_id)
        self._add_metrics_sheet(wb, fund_id)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    def _add_overview_sheet(self, wb: Workbook, fund: Fund):
        """Add fund overview sheet"""
        ws = wb.create_sheet("Overview", 0)

        # Title
        ws['A1'] = "Fund Overview"
        ws['A1'].font = Font(size=16, bold=True)

        # Fund details
        row = 3
        details = [
            ("Fund Name:", fund.name),
            ("GP Name:", fund.gp_name or "N/A"),
            ("Fund Type:", fund.fund_type or "N/A"),
            ("Vintage Year:", fund.vintage_year or "N/A"),
            ("Created:", fund.created_at.strftime("%Y-%m-%d") if fund.created_at else "N/A"),
        ]

        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Style
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _add_capital_calls_sheet(self, wb: Workbook, fund_id: int):
        """Add capital calls sheet"""
        ws = wb.create_sheet("Capital Calls")

        # Headers
        headers = ["Date", "Call Type", "Amount", "Description"]
        self._write_header_row(ws, headers, 1)

        # Get data
        capital_calls = self.db.query(CapitalCall).filter(
            CapitalCall.fund_id == fund_id
        ).order_by(CapitalCall.call_date).all()

        # Write data
        row = 2
        total = 0
        for call in capital_calls:
            ws[f'A{row}'] = call.call_date.strftime("%Y-%m-%d")
            ws[f'B{row}'] = call.call_type or ""
            ws[f'C{row}'] = float(call.amount)
            ws[f'C{row}'].number_format = '$#,##0.00'
            ws[f'D{row}'] = call.description or ""
            total += float(call.amount)
            row += 1

        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'] = total
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].fill = PatternFill(start_color="FFFF00", fill_type="solid")

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40

    def _add_distributions_sheet(self, wb: Workbook, fund_id: int):
        """Add distributions sheet"""
        ws = wb.create_sheet("Distributions")

        # Headers
        headers = ["Date", "Type", "Amount", "Recallable", "Description"]
        self._write_header_row(ws, headers, 1)

        # Get data
        distributions = self.db.query(Distribution).filter(
            Distribution.fund_id == fund_id
        ).order_by(Distribution.distribution_date).all()

        # Write data
        row = 2
        total = 0
        for dist in distributions:
            ws[f'A{row}'] = dist.distribution_date.strftime("%Y-%m-%d")
            ws[f'B{row}'] = dist.distribution_type or ""
            ws[f'C{row}'] = float(dist.amount)
            ws[f'C{row}'].number_format = '$#,##0.00'
            ws[f'D{row}'] = "Yes" if dist.is_recallable else "No"
            ws[f'E{row}'] = dist.description or ""
            total += float(dist.amount)
            row += 1

        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'] = total
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].fill = PatternFill(start_color="FFFF00", fill_type="solid")

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40

    def _add_adjustments_sheet(self, wb: Workbook, fund_id: int):
        """Add adjustments sheet"""
        ws = wb.create_sheet("Adjustments")

        # Headers
        headers = ["Date", "Type", "Category", "Amount", "Description"]
        self._write_header_row(ws, headers, 1)

        # Get data
        adjustments = self.db.query(Adjustment).filter(
            Adjustment.fund_id == fund_id
        ).order_by(Adjustment.adjustment_date).all()

        # Write data
        row = 2
        total = 0
        for adj in adjustments:
            ws[f'A{row}'] = adj.adjustment_date.strftime("%Y-%m-%d")
            ws[f'B{row}'] = adj.adjustment_type or ""
            ws[f'C{row}'] = adj.category or ""
            ws[f'D{row}'] = float(adj.amount)
            ws[f'D{row}'].number_format = '$#,##0.00'
            ws[f'E{row}'] = adj.description or ""
            total += float(adj.amount)
            row += 1

        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'D{row}'] = total
        ws[f'D{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40

    def _add_metrics_sheet(self, wb: Workbook, fund_id: int):
        """Add metrics summary sheet"""
        ws = wb.create_sheet("Metrics")

        # Title
        ws['A1'] = "Performance Metrics"
        ws['A1'].font = Font(size=14, bold=True)

        # Calculate metrics
        metrics = self.metrics_calculator.calculate_all_metrics(fund_id)

        # Metrics data
        row = 3
        metrics_data = [
            ("Paid-In Capital (PIC)", metrics.get("pic", 0), "$#,##0.00"),
            ("Total Distributions", metrics.get("total_distributions", 0), "$#,##0.00"),
            ("DPI (Distribution to Paid-In)", metrics.get("dpi", 0), "0.00"),
            ("IRR (Internal Rate of Return)", metrics.get("irr", 0), "0.00%"),
        ]

        for label, value, format_str in metrics_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = format_str
            ws[f'B{row}'].fill = PatternFill(start_color="E0F7FF", fill_type="solid")
            row += 1

        # Generated timestamp
        row += 2
        ws[f'A{row}'] = "Generated:"
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'B{row}'].font = Font(italic=True)

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _write_header_row(self, ws, headers: List[str], row: int):
        """Write styled header row"""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
